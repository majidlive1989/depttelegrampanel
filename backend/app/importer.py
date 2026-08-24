import io
import re
from openpyxl import load_workbook
from pypdf import PdfReader

NAME_KEYS = {"name", "customer", "customer name", "client", "نام", "نام مشتری", "مشتری", "طرف حساب"}
DEBT_KEYS = {"debt", "balance", "amount", "debt amount", "مانده", "مانده حساب", "بدهی", "مبلغ بدهی", "طلب"}
ID_KEYS = {"id", "customer id", "code", "customer code", "کد", "کد مشتری", "شناسه"}

PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")


def normalize_text(v):
    return str(v or "").translate(PERSIAN_DIGITS).strip()


def parse_money(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = normalize_text(v).replace(",", "").replace("٬", "").replace(" ", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return int(float(s))
    except Exception:
        return None


def find_col(headers, keys):
    for i, h in enumerate(headers):
        if normalize_text(h).lower() in keys:
            return i
    return None


def parse_excel(data: bytes):
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return [], ["فایل اکسل خالی است."]
    headers = [normalize_text(x).lower() for x in raw[0]]
    name_i, debt_i, id_i = find_col(headers, NAME_KEYS), find_col(headers, DEBT_KEYS), find_col(headers, ID_KEYS)
    warnings = []
    if name_i is None or debt_i is None:
        return [], ["ستون نام مشتری یا مبلغ بدهی پیدا نشد. نام ستون‌ها را مثل «نام مشتری» و «بدهی» قرار دهید."]
    result = []
    for idx, r in enumerate(raw[1:], start=2):
        name = normalize_text(r[name_i] if name_i < len(r) else "")
        debt = parse_money(r[debt_i] if debt_i < len(r) else None)
        ext = normalize_text(r[id_i] if id_i is not None and id_i < len(r) else "") or None
        if not name or debt is None:
            warnings.append(f"ردیف {idx} نادیده گرفته شد.")
            continue
        result.append({"external_id": ext, "name": name, "debt_amount": max(0, debt)})
    return result, warnings[:30]


def parse_pdf(data: bytes):
    reader = PdfReader(io.BytesIO(data))
    text = "\n".join((p.extract_text() or "") for p in reader.pages)
    result, warnings = [], []
    for line in text.splitlines():
        line = normalize_text(line)
        # Generic fallback: text name followed by a reasonably long numeric amount.
        m = re.match(r"^(.{2,80}?)\s+([0-9][0-9,٬]{3,})\s*(?:تومان|ریال)?\s*$", line)
        if not m:
            continue
        name = m.group(1).strip(" :-|\t")
        debt = parse_money(m.group(2))
        if name and debt is not None:
            result.append({"external_id": None, "name": name, "debt_amount": max(0, debt)})
    if not result:
        warnings.append("ساختار PDF به‌صورت خودکار تشخیص داده نشد. برای نسخه اول Excel ورودی مطمئن‌تر است.")
    return result, warnings


def parse_file(filename: str, data: bytes):
    name = filename.lower()
    if name.endswith(".xlsx"):
        return parse_excel(data), "xlsx"
    if name.endswith(".pdf"):
        return parse_pdf(data), "pdf"
    raise ValueError("فقط فایل XLSX یا PDF پشتیبانی می‌شود.")
